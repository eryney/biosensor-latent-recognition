"""Build the supplementary tables workbook from processed outputs."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
PROCESSED = DATA / "processed"
METADATA = DATA / "metadata"
ANALYSIS = ROOT / "analysis"
OUT = ROOT / "supplement"
XLSX = OUT / "Supplementary_Tables.xlsx"


LEAD_CURVES = [
    "cipro_iEscSnFR.csv",
    "cipro_v7.1.csv",
    "cipro_v7.csv",
    "cipro_v9.csv",
    "dehp_v4.8.1.2.csv",
    "dehp_L194.csv",
    "thiamine_iEscSnFR.csv",
    "ergothioneine_v9.csv",
    "carnitine_v9.csv",
]

SENSOR_META = {
    "v4.6": ("iNicSnFR1 (v4.6)", "nicotinic/cholinergic trunk"),
    "V4.8.1.2": ("Nic-1 intermediate (V4.8.1.2)", "nicotinic/cholinergic trunk"),
    "V6": ("Nic-2 intermediate (V6)", "nicotinic/cholinergic trunk"),
    "cc93": ("iNicSnFR3a (cc93)", "nicotinic/cholinergic trunk"),
    "V7": ("iNicSnFR3b (V7)", "nicotinic/cholinergic trunk"),
    "V7.1": ("ACh-1 intermediate (V7.1)", "acetylcholine branch"),
    "V7.1.2": ("ACh-2 intermediate (V7.1.2)", "acetylcholine branch"),
    "V8": ("ACh-3 intermediate (V8)", "acetylcholine branch"),
    "V9": ("iAChSnFR (V9)", "acetylcholine branch"),
    "L194": ("iSeroSnFR intermediate (L194)", "serotonin branch"),
    "AK1": ("AK1 intermediate", "opioid/ketamine branch"),
    "Fent2 436L": ("iFentanylSnFR negative control", "sequence outgroup negative control"),
    "iEscSnFR": ("iEscSnFR", "escitalopram branch"),
    "Tap1.0": ("iTapentadolSnFR (Tap1.0)", "tapentadol branch"),
    "iCytSnFR": ("iCytSnFR", "cytisine branch"),
    "iCytBrEtSnFR": ("iCyt_BrEt_SnFR", "cytisine analog branch"),
    "iFloxSnFR": ("iFluoxSnFR (iFloxSnFR)", "fluoxetine branch"),
    "iLevaphenolSnFR1.0": ("iLevorphanolSnFR1.0", "levorphanol branch"),
}


def fit_class(row: pd.Series) -> str:
    if bool(row.get("is_noise", False)):
        return "noise"
    if bool(row.get("is_step_fn", False)):
        return "step-like"
    if bool(row.get("success", False)) and row["ec50_uM"] < 400 and row["df_max"] < 20:
        return "within-range reliable"
    return "lower-bound or out-of-range"


def build_tables() -> dict[str, pd.DataFrame]:
    matrix = pd.read_csv(PROCESSED / "response_matrix.csv")
    sem = pd.read_csv(PROCESSED / "response_sem_matrix.csv")
    ligand_summary = pd.read_csv(ANALYSIS / "figure2_ligand_scope_summary.csv")
    fits = pd.read_csv(PROCESSED / "dose_response_fits.csv")
    exclusions = pd.read_csv(PROCESSED / "dose_response_exclusions.csv")

    fits["fit_class"] = fits.apply(fit_class, axis=1)
    fits = fits[
        [
            "file",
            "ligand",
            "sensor",
            "n_points",
            "baseline",
            "df_max",
            "ec50_uM",
            "hill_n",
            "r2",
            "lod_uM",
            "fit_class",
            "message",
            "source_file",
            "source_sheet",
            "excluded_wells",
        ]
    ]

    h2 = fits[fits["excluded_wells"].fillna("").str.contains("H2")][
        ["file", "source_file", "source_sheet", "excluded_wells"]
    ].copy()
    h2.insert(1, "sheet", h2.pop("source_sheet"))
    h2.insert(2, "output", h2.pop("file"))
    h2.insert(3, "reason", "bad ciprofloxacin H2 well; 200 uM point recomputed from remaining 2 replicates")
    h2 = h2.rename(columns={"source_file": "file"})
    h2 = h2[["file", "sheet", "output", "reason", "excluded_wells"]]
    curve_exclusions = exclusions.copy()
    curve_exclusions["excluded_wells"] = ""
    exclusions_full = pd.concat([curve_exclusions, h2], ignore_index=True)

    lead_rows = []
    for fname in LEAD_CURVES:
        curve = pd.read_csv(PROCESSED / "dose_responses" / fname)
        fit = fits.loc[fits["file"] == fname].iloc[0]
        curve.insert(0, "file", fname)
        curve.insert(1, "ligand", fit["ligand"])
        curve.insert(2, "sensor", fit["sensor"])
        lead_rows.append(curve)
    lead_curves = pd.concat(lead_rows, ignore_index=True)

    seq_source = pd.read_csv(METADATA / "sensor_sequences.csv")
    seq_rows = []
    for _seq_row in seq_source.itertuples(index=False):
        sensor = str(_seq_row.sensor)
        sequence = str(_seq_row.sequence).upper()
        display, lineage = SENSOR_META.get(sensor, (sensor, "unresolved"))
        seq_rows.append(
            {
                "sensor": sensor,
                "paper_facing_name": display,
                "lineage": lineage,
                "included_in_scaffold_sequence_analysis": sensor != "Fent2 436L",
                "sequence_length": len(sequence),
                "residue_357_sequence_index": sequence[356] if len(sequence) >= 357 else "",
                "residue_436_sequence_index": sequence[435] if len(sequence) >= 436 else "",
                "sequence": sequence,
            }
        )
    sensor_sequences = pd.DataFrame(seq_rows)

    readme = pd.DataFrame(
        {
            "field": [
                "source_screen_raw",
                "source_dose_raw",
                "screen_matrix",
                "screen_thresholds",
                "dose_curve_exclusions",
                "ciprofloxacin_H2_rule",
                "fit_class_rule",
                "sensor_sequence_table",
                "generated_by",
            ],
            "value": [
                "raw_screening_data",
                "raw_dose_response_data",
                "Full-precision response and SEM matrices rebuilt from per-sensor workbooks.",
                "Permissive hit: dF/F0 > 0.3; strong hit: dF/F0 > 1.0.",
                "MEHP_cc93, aspartame_iLevaSnFR, bilirubin_V4.8.1.2, theobromine_Tap1.0.",
                "For ciprofloxacin dose responses, H2 was excluded and the 200 uM point was recomputed from the remaining 2 replicates.",
                "within-range reliable means EC50 < 400 uM and dFmax < 20; otherwise lower-bound/out-of-range unless flagged as noise or step-like.",
                "S7_Sensor_Sequences reports each sequence used in the screen plus lineage labels and key sequence-index residues.",
                "analysis/build_supplement_tables.py",
            ],
        }
    )

    return {
        "README": readme,
        "S1_Response_Matrix": matrix,
        "S2_Response_SEM": sem,
        "S3_Ligand_Summary": ligand_summary,
        "S4_Dose_Response_Fits": fits,
        "S5_Exclusions_Notes": exclusions_full,
        "S6_Fig3_Curve_Points": lead_curves,
        "S7_Sensor_Sequences": sensor_sequences,
    }


def write_csvs(tables: dict[str, pd.DataFrame]) -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    for name, df in tables.items():
        df.to_csv(OUT / f"{name}.csv", index=False)


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(bold=True, color="1F4E78")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=False)
        for idx, column_cells in enumerate(ws.columns, start=1):
            header = str(column_cells[0].value or "")
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells[:250])
            width = min(max(max_len + 2, len(header) + 2, 10), 42)
            if header in {"SMILES", "source_file", "message", "value", "reason"}:
                width = min(max(width, 28), 60)
            ws.column_dimensions[get_column_letter(idx)].width = width
        if ws.title == "README":
            ws["A1"].font = title_font
            ws.column_dimensions["A"].width = 28
            ws.column_dimensions["B"].width = 95
            for row in ws.iter_rows(min_row=2, max_col=2):
                row[1].alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(path)


def build() -> None:
    tables = build_tables()
    write_csvs(tables)
    with pd.ExcelWriter(XLSX, engine="openpyxl") as writer:
        for name, df in tables.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)
    style_workbook(XLSX)
    print(XLSX)


if __name__ == "__main__":
    build()
