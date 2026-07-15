"""Rebuild processed data tables from the raw plate-reader workbooks."""

from __future__ import annotations

import math
import re
import shutil
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
RAW = DATA / "raw"
SCREEN_RAW = RAW / "raw_screening_data"
DOSE_RAW = RAW / "raw_dose_response_data"
PROCESSED = DATA / "processed"
METADATA = DATA / "metadata"
DOSE_OUT = PROCESSED / "dose_responses"

sys.path.insert(0, str(ROOT / "scripts"))
from utils import compute_lod, fit_hill  # noqa: E402


SENSOR_WORKBOOKS = {
    "AK1": "ak1 screen_analyzed.xlsx",
    "cc93": "cc93_screen_analyzed.xlsx",
    "Fent2 436L": "Fent2.436L_screen_analyzed.xlsx",
    "iCytBrEtSnFR": "iCytBrEtSnFR_screen_analyzed.xlsx",
    "iCytSnFR": "iCytSnFR_screen_analyzed.xlsx",
    "iEscSnFR": "iEscSnFR_screen_analyzed.xlsx",
    "iFloxSnFR": "iFloxSnFR_screen_analyzed.xlsx",
    "iLevaphenolSnFR1.0": "iLevaSnFR1.0_screen_analyzed.xlsx",
    "L194": "L194_screen_analyzed.xlsx",
    "Tap1.0": "Tap1.0_screen_analyzed.xlsx",
    "V4.8.1.2": "v4.8.1.2_screen_analyzed.xlsx",
    "v4.6": "v4.6_screen_analyzed.xlsx",
    "V6": "v6_screen_analyzed.xlsx",
    "V7": "v7_screen_analyzed.xlsx",
    "V7.1": "v7.1_screen_analyzed.xlsx",
    "V7.1.2": "v7.1.2_screen_analyzed.xlsx",
    "V8": "v8 screen_analyzed.xlsx",
    "V9": "v9 screen_analyzed.xlsx",
}

SENSOR_ALIASES = {
    "ak1": "ak1",
    "ak2": "ak2",
    "cc93": "cc93",
    "iescsnfr": "iEscSnFR",
    "iesc": "iEscSnFR",
    "icytbrsnfr": "iCytBrSnFR",
    "ilevasnfr": "iLevaSnFR",
    "ilevasnfr1.0": "iLevaSnFR",
    "l194": "L194",
    "tap1.0": "tap1.0",
    "tap10": "tap1.0",
    "v4.8.1.2": "v4.8.1.2",
    "v6": "v6",
    "v7": "v7",
    "v7.1": "v7.1",
    "v7.1.2": "v7.1.2",
    "v8": "v8",
    "v9": "v9",
}

LIGAND_ALIASES = {
    "mehp": "mehp",
    "estradio": "estradiol",
    "beta-estradio": "estradiol",
    "beta-estradiol": "estradiol",
    "caffeine": "caffeine",
    "cipro": "cipro",
    "deet": "deet",
    "thiamine": "thiamine",
    "carnitine": "carnitine",
    "atenolol": "atenolol",
    "dehp": "dehp",
    "fipronil": "fipronil",
    "aspartame": "aspartame",
    "atrazine": "atrazine",
    "beta-hydroxybutyrate": "betahydroxybutyrate",
    "milirone": "milirone",
    "bilirubin": "bilirubin",
    "cyclosporin a": "cyclosporinA",
    "tyramine hydrochloride": "tyramine",
    "ergothioneine": "ergothioneine",
    "ergothionein": "ergothioneine",
    "folic acid": "folic_acid",
    "digoxin": "digoxin",
    "theobromine": "theobromine",
    "tryptamine": "tryptamine",
    "epi": "epinephrine",
    "l-thyroxine": "l-thyroxine",
    "thyroxine": "l-thyroxine",
}

EXCLUDED_CURVES = {
    ("mehp", "cc93"),
    ("aspartame", "iLevaSnFR"),
    ("bilirubin", "v4.8.1.2"),
    ("theobromine", "tap1.0"),
}

CIPRO_H2_EXCLUSIONS = {
    "20240726_ciprofloxacin_V7.1_3xpbsph7.xlsx",
    "20240726_thiamine_ciprofloxacin_V7_3xpbsph7.xlsx",
    "20240726_thiamine_ciprofloxacin_V9_3xpbsph7.xlsx",
    "20240726_thiamine_ciprofloxacin_iEscSnFR_3xpbsph7.xlsx",
}


def safe_float(value) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return np.nan


def clean_ligand_name(name: str) -> str:
    return re.sub(r"\s+", " ", str(name).strip())


def rebuild_screen() -> None:
    analyzed = SCREEN_RAW / "Analayzed_Data"
    combined = pd.read_excel(analyzed / "2024Summer_Combined_Heatmap.xlsx")
    combined = combined.rename(columns={combined.columns[0]: "ligand"}).dropna(subset=["ligand"])
    ligands = [clean_ligand_name(x) for x in combined["ligand"].tolist()]

    stacks: dict[str, list[float]] = {}
    sem_stacks: dict[str, list[float]] = {}

    for sensor, workbook_name in SENSOR_WORKBOOKS.items():
        wb = load_workbook(analyzed / workbook_name, data_only=True, read_only=False)
        ws = wb["Sheet2"]
        values = []
        for col in range(2, 13):
            for row in range(20, 28):
                values.append(safe_float(ws.cell(row=row, column=col).value))
        stacks[sensor] = values

        # Propagate the per-well average SEM against its matched solvent baseline.
        # This is mainly for audit/supplemental reporting; the main heatmap uses dF/F0.
        sems: list[float] = []
        ws1 = wb["Sheet1"]
        average_row = None
        for row in range(1, ws1.max_row + 1):
            if ws1.cell(row=row, column=1).value == "Average":
                average_row = row
                break
        if average_row is None:
            raise ValueError(f"Could not find Average block in {workbook_name}")

        for idx in range(88):
            avg = safe_float(ws.cell(row=10 + (idx % 8), column=2 + (idx // 8)).value)
            baseline = safe_float(ws.cell(row=30 + (idx % 8), column=2 + (idx // 8)).value)
            baseline_sem = np.nan
            # Sheet1 stores raw SEM beside the average block, columns N:X.
            raw_sem = safe_float(ws1.cell(row=average_row + 2 + (idx % 8), column=14 + (idx // 8)).value)
            for row in range(2, 8):
                if safe_float(ws.cell(row=row, column=2).value) == baseline:
                    baseline_sem = safe_float(ws.cell(row=row, column=3).value)
                    break
            if np.isfinite(avg) and np.isfinite(baseline) and baseline != 0 and np.isfinite(raw_sem) and np.isfinite(baseline_sem):
                sems.append(float(abs(math.sqrt(raw_sem**2 + baseline_sem**2) / baseline)))
            else:
                sems.append(np.nan)
        sem_stacks[sensor] = sems

    selected_indices: list[int] = []
    cursor = 0
    for _, row in combined.iterrows():
        target = np.array([row[sensor] for sensor in SENSOR_WORKBOOKS], dtype=float)
        found = None
        for idx in range(cursor, 88):
            candidate = np.array([stacks[sensor][idx] for sensor in SENSOR_WORKBOOKS], dtype=float)
            if np.all(np.isfinite(candidate)) and np.allclose(np.round(candidate, 1), target, atol=1e-12):
                found = idx
                break
        if found is None:
            ligand = row["ligand"]
            raise ValueError(f"Could not map rounded combined heatmap row back to full precision stack: {ligand}")
        selected_indices.append(found)
        cursor = found + 1

    matrix = pd.DataFrame(
        {sensor: [stacks[sensor][idx] for idx in selected_indices] for sensor in SENSOR_WORKBOOKS},
        index=ligands,
    )
    sem_matrix = pd.DataFrame(
        {sensor: [sem_stacks[sensor][idx] for idx in selected_indices] for sensor in SENSOR_WORKBOOKS},
        index=ligands,
    )
    matrix = matrix.drop(index=["nothing"], errors="ignore")
    sem_matrix = sem_matrix.drop(index=["nothing"], errors="ignore")
    matrix.index.name = "ligand"
    sem_matrix.index.name = "ligand"

    matrix.to_csv(PROCESSED / "response_matrix.csv")
    sem_matrix.to_csv(PROCESSED / "response_sem_matrix.csv")

    for fname in ["ligand_smiles.csv", "ligand_categories.csv"]:
        shutil.copy2(METADATA / fname, PROCESSED / fname)


def parse_sensor_from_filename(path: Path) -> str:
    stem = path.stem.replace("_3xpbsph7", "")
    raw = stem.split("_")[-1].lower()
    return SENSOR_ALIASES.get(raw, raw)


def canonical_ligand(sheet_name: str) -> str:
    key = sheet_name.strip().lower()
    return LIGAND_ALIASES.get(key, key.replace(" ", "_"))


def curve_from_sheet(path: Path, sheet_name: str) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name]

    raw = np.array(
        [[safe_float(ws.cell(row=r, column=c).value) for c in range(1, 10)] for r in range(1, 4)],
        dtype=float,
    )
    ligand = canonical_ligand(sheet_name)

    baseline_vals = raw[:, 0]
    baseline_mean = float(np.nanmean(baseline_vals))
    baseline_sd = float(np.nanstd(baseline_vals, ddof=1))

    concentrations = [200.0]
    for _ in range(7):
        concentrations.append(concentrations[-1] / math.sqrt(10))

    rows = []
    for j, conc in enumerate(concentrations, start=1):
        vals = raw[:, j].astype(float)
        n = int(np.count_nonzero(np.isfinite(vals)))
        excluded = False
        if ligand == "cipro" and path.name in CIPRO_H2_EXCLUSIONS and j == 1:
            vals[2] = np.nan
            n = int(np.count_nonzero(np.isfinite(vals)))
            excluded = True
        mean = float(np.nanmean(vals))
        sd = float(np.nanstd(vals, ddof=1)) if n > 1 else 0.0
        diff = mean - baseline_mean
        dff = diff / baseline_mean
        if diff != 0 and baseline_mean != 0:
            propagated = (((math.sqrt(sd**2 + baseline_sd**2) / diff) ** 2 + (baseline_sd / baseline_mean) ** 2) * dff) / math.sqrt(n)
            sem = abs(float(propagated))
        else:
            sem = np.nan
        rows.append(
            {
                "conc_uM": conc,
                "dF_F0": dff,
                "SEM": sem,
                "n": n,
                "excluded_wells": "H2" if excluded else "",
                "source_file": path.name,
                "source_sheet": sheet_name,
            }
        )

    return pd.DataFrame(rows).sort_values("conc_uM").reset_index(drop=True)


def rebuild_dose_responses() -> None:
    rows = []
    exclusions = []
    for path in sorted(DOSE_RAW.glob("*.xlsx")):
        sensor = parse_sensor_from_filename(path)
        wb = load_workbook(path, data_only=True, read_only=True)
        for sheet_name in [s for s in wb.sheetnames if s != "Result sheet"]:
            ligand = canonical_ligand(sheet_name)
            out_stem = f"{ligand}_{sensor}"
            if (ligand, sensor) in EXCLUDED_CURVES:
                exclusions.append({"file": path.name, "sheet": sheet_name, "output": out_stem + ".csv", "reason": "excluded weak/negative extra curve"})
                continue

            curve = curve_from_sheet(path, sheet_name)
            out_path = DOSE_OUT / f"{out_stem}.csv"
            curve[["conc_uM", "dF_F0", "SEM"]].to_csv(out_path, index=False)

            conc = curve["conc_uM"].to_numpy()
            signal = curve["dF_F0"].to_numpy()
            sem = curve["SEM"].to_numpy()
            fit = fit_hill(conc, signal, sem)
            lod = compute_lod(conc, fit, sem) if fit["success"] else np.nan
            rows.append(
                {
                    "file": out_path.name,
                    "ligand": ligand,
                    "sensor": sensor,
                    "n_points": len(curve),
                    "baseline": fit["baseline"],
                    "df_max": fit["df_max"],
                    "ec50_uM": fit["ec50"],
                    "hill_n": fit["n"],
                    "r2": fit["r2"],
                    "lod_uM": lod,
                    "success": fit["success"],
                    "is_noise": fit.get("is_noise", False),
                    "is_step_fn": fit.get("is_step_fn", False),
                    "message": fit["message"],
                    "source_file": path.name,
                    "source_sheet": sheet_name,
                    "excluded_wells": ";".join(sorted(set(x for x in curve["excluded_wells"] if x))),
                }
            )

    fits = pd.DataFrame(rows).sort_values(["ligand", "sensor"]).reset_index(drop=True)
    fits.to_csv(PROCESSED / "dose_response_fits.csv", index=False)
    pd.DataFrame(exclusions).to_csv(PROCESSED / "dose_response_exclusions.csv", index=False)


def main() -> None:
    PROCESSED.mkdir(parents=True, exist_ok=True)
    DOSE_OUT.mkdir(parents=True, exist_ok=True)
    for old in DOSE_OUT.glob("*.csv"):
        old.unlink()

    rebuild_screen()
    rebuild_dose_responses()

    matrix = pd.read_csv(PROCESSED / "response_matrix.csv", index_col=0)
    fits = pd.read_csv(PROCESSED / "dose_response_fits.csv")
    print(f"Wrote {matrix.shape[0]} x {matrix.shape[1]} screen matrix to {PROCESSED / 'response_matrix.csv'}")
    print(f"Wrote {len(fits)} dose-response fits to {PROCESSED / 'dose_response_fits.csv'}")


if __name__ == "__main__":
    main()
